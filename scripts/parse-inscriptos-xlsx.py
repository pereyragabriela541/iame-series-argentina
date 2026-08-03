#!/usr/bin/env python3
import json
import sys
import unicodedata
from pathlib import Path

import openpyxl

SHEET_TO_SLUG = {
    "MINI": "60-mini",
    "OKN JUNIOR": "okn-junior",
    "OKN": "okn",
    "JUNIOR OLD": "junior",
    "SENIOR OLD": "senior",
    "MASTER OLD": "master-my10",
    "HONDA 390": "senior-pro-390-honda",
    "ACADEMY": "academy",
}


def norm(s):
    if not s:
        return ""
    t = str(s).upper().replace(",", " ")
    t = "".join(
        c for c in unicodedata.normalize("NFD", t) if unicodedata.category(c) != "Mn"
    )
    return " ".join(t.split())


def main():
    path = Path(sys.argv[1])
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out = []

    for sn in wb.sheetnames:
        ws = wb[sn]
        rows = list(ws.iter_rows(values_only=True))
        name_col = kart_col = None
        start = 0

        for i, row in enumerate(rows):
            cells = [str(c).upper() if c is not None else "" for c in row]
            if "PILOTO" in cells:
                start = i + 1
                name_col = cells.index("PILOTO")
                kart_col = cells.index("NRO KART") if "NRO KART" in cells else None
                break

        if kart_col is None or name_col is None:
            continue

        if sn == "OKN JUNIOR" and len(rows) > 1:
            hdr = [str(c).upper() if c is not None else "" for c in rows[1]]
            if len(hdr) > 2 and hdr[1] == "NRO KART" and hdr[2] == "PILOTO":
                name_col, kart_col, start = 2, 1, 2

        for row in rows[start:]:
            if not row:
                continue
            name = norm(row[name_col] if name_col < len(row) else None)
            kart = row[kart_col] if kart_col < len(row) else None
            if not name or kart in (None, ""):
                continue
            try:
                kart = str(int(float(kart)))
            except (TypeError, ValueError):
                continue
            out.append(
                {
                    "sheet": sn,
                    "slug": SHEET_TO_SLUG.get(sn, sn),
                    "name": name,
                    "kart": kart,
                }
            )

    print(json.dumps(out))


if __name__ == "__main__":
    main()
